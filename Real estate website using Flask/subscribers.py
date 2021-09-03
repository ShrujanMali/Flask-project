from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import smtplib

def createexcel(email):
    """ If Subscribers xlsx file is dose not exist then it creates new one and save email id into it"""
    headers = ['Emails']
    workbook_name = 'SubscribersList.xlsx'
    wb = Workbook()
    page = wb.active
    page.title = 'Emails'
    page.append(headers) # write the headers to the first line

    # Data to write:
    page.append([email])
    wb.save(filename = workbook_name)
    thanksforsub(email)


def emailappend(email):
    """ This function executes when the user enter email on home page for subscrition then it takes mail id and append into existing xlsx file """
    workbook_name = 'SubscribersList.xlsx'
    wb = load_workbook(workbook_name)
    page = wb.active
    page.append([email])
    wb.save(filename=workbook_name)
    thanksforsub(email)

def thanksforsub(newmail):
    """ This function send welcome mail to user when they subscribe """
    message = "Thank for subscribing us"
    server = smtplib.SMTP("smtp.gmail.com", 587)
    server.starttls()
    server.login("Email address", "password")            # Put email address and password
    server.sendmail("Email address", newmail, message)
